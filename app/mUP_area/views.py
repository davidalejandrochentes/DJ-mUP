from django.shortcuts import render, redirect, get_object_or_404
from .models import Area, MantenimientoArea, TipoMantenimientoArea, DiasParaAlerta
from .forms import AreaForm, MantenimientoAreaForm, DiasParaAlertaForm
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseRedirect

import openpyxl
from openpyxl.styles import Font, PatternFill



# vistas generales -----------------------------------------------

@login_required
def area(request):
    dias_alerta = DiasParaAlerta.objects.first().días
    search_query = request.GET.get('search', '')
    areas = Area.objects.filter(nombre__icontains=search_query)
    alertas = [{'area': area, 'dias_restantes': area.dias_restantes_mantenimiento()} for area in Area.objects.all() if area.dias_restantes_mantenimiento() <= dias_alerta]
    alertas_ordenadas = sorted(alertas, key=lambda x: x['dias_restantes'])
    context = {'areas': areas, 'total_areas': areas.count(), 'alertas': alertas_ordenadas, 'total_alertas': len(alertas_ordenadas)}
    return render(request, 'mUP_area/area.html', context)

@login_required
def alertas(request):
    dias_alert = get_object_or_404(DiasParaAlerta, id=1)
    if request.method == 'POST':
        alert_form = DiasParaAlertaForm(request.POST, instance=dias_alert)
        if alert_form.is_valid():
            if alert_form.cleaned_data.get('días') >= 1:
                alert_form.save()
            return redirect('area_alertas')
    else:
        alert_form = DiasParaAlertaForm(instance=dias_alert)
    dias_alerta = dias_alert.días
    alertas = [{'area': area, 'dias_restantes': area.dias_restantes_mantenimiento()} for area in Area.objects.filter(nombre__icontains=request.GET.get('search', '')) if area.dias_restantes_mantenimiento() <= dias_alerta]
    alertas_ordenadas = sorted(alertas, key=lambda x: x['dias_restantes'])
    context = {'alertas': alertas_ordenadas, 'total_alertas': len(alertas_ordenadas), 'alert_form': alert_form}
    return render(request, 'mUP_area/alertas.html', context)

@login_required
def tabla_mantenimientos(request):
    areas = Area.objects.prefetch_related('mantenimientoarea_set').all()
    for area in areas:
        area.mantenimientos = area.mantenimientoarea_set.all().order_by('-fecha_fin', '-hora_fin')
    context = {'areas': areas, 'tipos_mantenimiento': TipoMantenimientoArea.objects.all()}
    return render(request, 'mUP_area/tablas.html', context)

@login_required
def crear_area(request):
    form = AreaForm(request.POST or None, request.FILES or None)
    if request.method == 'POST':
        if form.is_valid():
            intervalo_mantenimiento = form.cleaned_data.get('intervalo_mantenimiento')
            if intervalo_mantenimiento < 0:
                form.add_error('intervalo_mantenimiento', 'El intervalo de mantenimiento no puede ser un número negativo')
            else:
                form.save()
                return redirect('area')
        else:
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
    return render(request, 'mUP_area/nueva.html', {'form': form})

@login_required
def detalles(request, id):
    area = get_object_or_404(Area, id=id)
    mantenimientos = area.mantenimientoarea_set.all().order_by('-fecha_fin', '-hora_fin')
    form = AreaForm(request.POST or None, request.FILES or None, instance=area)
    if request.method == 'POST':
        if form.is_valid():
            intervalo_mantenimiento = form.cleaned_data.get('intervalo_mantenimiento')
            if intervalo_mantenimiento < 0:
                form.add_error('intervalo_mantenimiento', 'El intervalo de mantenimiento no puede ser un número negativo')
                return render(request, 'mUP_area/detalles.html', {
                    'area': area,
                    'form': form,
                    'id': id,
                    'mantenimientos': mantenimientos,
                })
            else:
                form.save()
                return redirect('detalles_area', id=id)
        else:
            return render(request, 'mUP_area/detalles.html', {
                'area': area,
                'form': form,
                'id': id,
                'mantenimientos': mantenimientos,
            })
    return render(request, 'mUP_area/detalles.html', {
        'area': area,
        'form': form,
        'id': id,
        'mantenimientos': mantenimientos,
    })

@login_required
def eliminar(request, id):
    area = get_object_or_404(Area, id = id)
    area.delete()
    return redirect ('area') 

# fin de vistas generales-----------------------------------------------    


@login_required
def eliminar_mantenimiento(request, id):
    mantenimiento = get_object_or_404(MantenimientoArea, id=id)
    mantenimiento.delete()
    previous_url = request.META.get('HTTP_REFERER')
    return HttpResponseRedirect(previous_url)
    
@login_required
def mantenimientos_area(request, id, mant):
    area = get_object_or_404(Area, id=id)
    tipo_mantenimiento = get_object_or_404(TipoMantenimientoArea, id=mant)
    mantenimientos = area.mantenimientoarea_set.filter(tipo=tipo_mantenimiento).order_by('-fecha_fin', '-hora_fin')
    context = {
        'area': area,
        'tipo_mantenimiento': tipo_mantenimiento,
        'mantenimientos': mantenimientos,
    }
    return render(request, 'mUP_area/manteniminetos.html', context)    

@login_required
def mod_mantenimiento_area(request, id, mant):
    mantenimiento = get_object_or_404(MantenimientoArea, id=id)
    tipo_mantenimiento = get_object_or_404(TipoMantenimientoArea, id=mant)
    area = mantenimiento.area
    if request.method == 'POST':
        form_mant = MantenimientoAreaForm(request.POST, request.FILES, instance=mantenimiento)
        if form_mant.is_valid():
            form_mant.save()
            return redirect('mantenimientos_area', id=area.id, mant=mant)
        messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
    else:
        form_mant = MantenimientoAreaForm(instance=mantenimiento)
    context = {
        'form_mant': form_mant,
        'area': area,
        'tipo_mantenimiento': tipo_mantenimiento,
    }
    return render(request, 'mUP_area/mod_mantenimineto.html', context)

@login_required
def nuevo_mantenimiento_area(request, id, mant):
    area = get_object_or_404(Area, id=id)
    tipo_mantenimiento = get_object_or_404(TipoMantenimientoArea, id=mant)
    if request.method == 'POST':
        form_mant = MantenimientoAreaForm(request.POST, request.FILES)
        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.area = area
            mantenimiento.tipo = tipo_mantenimiento
            mantenimiento.save()
            return redirect('mantenimientos_area', id=area.id, mant=mant)
        messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
    else:
        form_mant = MantenimientoAreaForm()
    context = {
        'form_mant': form_mant,
        'area': area,
        'tipo_mantenimiento': tipo_mantenimiento,
    }
    return render(request, 'mUP_area/nuevo_mantenimineto.html', context)

#---------------------------------------------------------------------------------

@login_required
def documento_general_mantenimientos_area(request):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    tipo_mantenimiento_id = request.GET.get('tipo_mantenimiento')

    mantenimientos = MantenimientoArea.objects.filter(fecha_fin__year=anio)

    if mes:
        mantenimientos = mantenimientos.filter(fecha_fin__month=mes)

    if tipo_mantenimiento_id:  # Si se seleccionó un tipo de mantenimiento
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoArea, pk=tipo_mantenimiento_id)
        mantenimientos = mantenimientos.filter(tipo=tipo_mantenimiento)

    mantenimientos = mantenimientos.order_by('-fecha_fin', '-hora_fin')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if mes:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_areas_{}_{}.xlsx"'.format(mes, anio)
    else:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_areas_{}.xlsx"'.format(anio)

    wb = openpyxl.Workbook()
    ws = wb.active

    headers = ['Area', 'Tipo', 'Fecha I', 'Hora I', 'Fecha F', 'Hora F', 'Operador', 'Descripción']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    row = 2
    for mantenimiento in mantenimientos:
        ws.append([
            mantenimiento.area.nombre,
            mantenimiento.tipo.tipo,
            mantenimiento.fecha_inicio,
            mantenimiento.hora_inicio,
            mantenimiento.fecha_fin,
            mantenimiento.hora_fin,
            mantenimiento.operador,
            mantenimiento.descripción
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response 




@login_required
def documento_mantenimientos_area(request, id, mant):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    tipo_mantenimiento_id = mant
    tipo_mantenimiento = get_object_or_404(TipoMantenimientoArea, id=mant) 

    area = get_object_or_404(Area, pk=id)
    mantenimientos = MantenimientoArea.objects.filter(area=area).order_by('-fecha_fin', '-hora_fin')

    if mes:
        mantenimientos = mantenimientos.filter(fecha_fin__month=mes)
    if anio:
        mantenimientos = mantenimientos.filter(fecha_fin__year=anio)
    if tipo_mantenimiento_id: # Si se seleccionó un tipo de mantenimiento
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoArea, pk=tipo_mantenimiento_id)
        mantenimientos = mantenimientos.filter(tipo=tipo_mantenimiento)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if mes:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_{}_de_{}_{}_{}.xlsx"'.format(tipo_mantenimiento.tipo, area.nombre, mes, anio)
    else:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_{}_de_{}_{}.xlsx"'.format(tipo_mantenimiento.tipo, area.nombre, anio)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Define los encabezados de la tabla
    headers = ['Operador', 'Fecha I', 'Hora I', 'Fecha F', 'Hora F', 'Descripción']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    # Agrega los datos de los mantenimientos
    row = 2
    for mantenimiento in mantenimientos:
        ws.append([
            mantenimiento.operador,
            mantenimiento.fecha_inicio,
            mantenimiento.hora_inicio,
            mantenimiento.fecha_fin,
            mantenimiento.hora_fin,
            mantenimiento.descripción
        ])

    # Ajusta el ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response