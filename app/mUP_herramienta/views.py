from django.shortcuts import render, redirect, get_object_or_404
from .models import Herramienta, MantenimientoHerramienta, TipoMantenimientoHerramienta
from .forms import HerramientaForm, MantenimientoHerramientaForm
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseRedirect
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage

import openpyxl
from openpyxl.styles import Font, PatternFill

# Create your views here.
@login_required
def herramienta(request):
    search_term = request.GET.get('search', '')
    herramientas_list = Herramienta.objects.filter(nombre__icontains=search_term)
    
    alertas = [
        {'herramienta': h, 'dias_restantes': h.dias_restantes_mantenimiento()}
        for h in Herramienta.objects.all() if h.dias_restantes_mantenimiento() <= 7
    ]
    alertas_ordenadas = sorted(alertas, key=lambda x: x['dias_restantes'])
    
    # Implementar paginación
    page = request.GET.get('page', 1)
    paginator = Paginator(herramientas_list, 10)  # 10 herramientas por página
    try:
        herramientas = paginator.page(page)
    except PageNotAnInteger:
        herramientas = paginator.page(1)
    except EmptyPage:
        herramientas = paginator.page(paginator.num_pages)
    
    context = {
        'herramientas': herramientas,
        'total_herramientas': herramientas_list.count(),
        'alertas': alertas_ordenadas,
        'total_alertas': len(alertas_ordenadas),
    }
    return render(request, 'mUP_herramienta/herramienta.html', context)

@login_required
def alertas(request):
    search_term = request.GET.get('search', '')
    alertas = [
        {'herramienta': h, 'dias_restantes': h.dias_restantes_mantenimiento()}
        for h in Herramienta.objects.filter(nombre__icontains=search_term)
        if h.dias_restantes_mantenimiento() <= 7
    ]
    alertas_ordenadas = sorted(alertas, key=lambda x: x['dias_restantes'])
    context = {
        'alertas': alertas_ordenadas,
        'total_alertas': len(alertas_ordenadas),
    }
    return render(request, 'mUP_herramienta/alertas.html', context)

@login_required
def tabla_mantenimientos(request):
    herramientas = Herramienta.objects.all()
    tipos_mantenimiento = TipoMantenimientoHerramienta.objects.all()
    for herramienta in herramientas:
        herramienta.mantenimientos = herramienta.mantenimientoherramienta_set.all().order_by('-fecha', '-hora')
    context = {
        'herramientas': herramientas,
        'tipos_mantenimiento': tipos_mantenimiento,
    }
    return render(request, 'mUP_herramienta/tablas.html', context)    


@login_required
def crear_herramienta(request):
    form = HerramientaForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('herramienta')
    if request.method == 'POST':
        messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
    context = {'form': form}
    return render(request, 'mUP_herramienta/nueva.html', context)           

@login_required
def eliminar(request, id):
    herramienta = get_object_or_404(Herramienta, id = id)
    herramienta.delete()
    return redirect ('herramienta') 

@login_required
def eliminar_mantenimiento(request, id):
    mantenimiento = get_object_or_404(MantenimientoHerramienta, id=id)
    mantenimiento.delete()
    previous_url = request.META.get('HTTP_REFERER')
    return HttpResponseRedirect(previous_url)    

from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseRedirect
from .forms import HerramientaForm, MantenimientoHerramientaForm
from .models import Herramienta, TipoMantenimientoHerramienta

@login_required
def detalles(request, id):
    herramienta = get_object_or_404(Herramienta, id=id)
    mantenimientos = herramienta.mantenimientoherramienta_set.all().order_by('-fecha', '-hora')
    tipos_mantenimiento = TipoMantenimientoHerramienta.objects.all()
    form = HerramientaForm(request.POST or None, request.FILES or None, instance=herramienta)
    form_mant = MantenimientoHerramientaForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return redirect('detalles_herramienta', id=id)
        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.herramienta = herramienta
            mantenimiento.save()
            return redirect('detalles_herramienta', id=id)
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    context = {
        'herramienta': herramienta,
        'form': form,
        'id': id,
        'form_mant': form_mant,
        'mantenimientos': mantenimientos,
        'tipos_mantenimiento': tipos_mantenimiento,
    }
    return render(request, 'mUP_herramienta/detalles.html', context)



def descargar_herramientas(request):
    herramientas = Herramienta.objects.all()

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="herramientas.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active

    headers = ['Nombre', 'Número de serie', 'Encargado', 'Teléfono encargado', 'Descripción', 'Fecha de adquisición', 'Costo', 'Proveedor', 'Ubicación', 'Estado de la herramienta', 'Fecha último mantenimiento', 'Intervalo mantenimiento']
    
    # Configuración de estilos para la cabecera
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
    
    # Inserción de datos
    for row, herramienta in enumerate(herramientas, start=2):
        ws.cell(row=row, column=1, value=herramienta.nombre)
        ws.cell(row=row, column=2, value=herramienta.número_de_serie)
        ws.cell(row=row, column=3, value=herramienta.encargado)
        ws.cell(row=row, column=4, value=herramienta.teléfono_encargado)
        ws.cell(row=row, column=5, value=herramienta.descripción)
        ws.cell(row=row, column=6, value=herramienta.fecha_de_adquisición)
        ws.cell(row=row, column=7, value=herramienta.costo)
        ws.cell(row=row, column=8, value=herramienta.proveedor)
        ws.cell(row=row, column=9, value=herramienta.ubicación)
        ws.cell(row=row, column=10, value=herramienta.estado_de_la_herramienta)
        ws.cell(row=row, column=11, value=herramienta.fecha_ultimo_mantenimiento)
        ws.cell(row=row, column=12, value=herramienta.intervalo_mantenimiento)

    # Ajuste del ancho de las columnas
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    wb.save(response)
    return response
