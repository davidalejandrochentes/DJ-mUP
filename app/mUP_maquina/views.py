from django.shortcuts import render, redirect, get_object_or_404
from .models import Maquina, MantenimientoMaquina, TipoMantenimientoMaquina, HorasParaAlerta
from .forms import MaquinaForm, MantenimientoMaquinaCorrectivoForm, MantenimientoMaquinaPreventivoForm, HorasParaAlertaForm
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse, HttpResponseRedirect

import openpyxl
from openpyxl.styles import Font, PatternFill



# vistas generales ----------------------------------------------------------------------------

@login_required
def maquina(request):
    horas_alerta = HorasParaAlerta.objects.values_list('horas', flat=True).first()
    search_query = request.GET.get('search', '')
    maquinas = Maquina.objects.filter(nombre__icontains=search_query)
    total_maquinas = maquinas.count()
    alertas = [
        {'maquina': maquina, 'horas_restantes': maquina.horas_restantes_mantenimiento()}
        for maquina in Maquina.objects.all()
        if maquina.horas_restantes_mantenimiento() <= horas_alerta
    ]
    alertas_ordenadas = sorted(alertas, key=lambda x: x['horas_restantes'])
    context = {
        'maquinas': maquinas,
        'total_maquinas': total_maquinas,
        'alertas': alertas_ordenadas,
        'total_alertas': len(alertas_ordenadas),
    }
    return render(request, 'mUP_maquina/maquina.html', context)

@login_required
def alertas(request):
    horas_alert = get_object_or_404(HorasParaAlerta, id=1)
    alert_form = HorasParaAlertaForm(request.POST or None, instance=horas_alert)
    if request.method == 'POST' and alert_form.is_valid():
        horas = alert_form.cleaned_data['horas']
        if horas >= 1:
            alert_form.save()
        return redirect('maquina_alertas')
    horas_alerta = horas_alert.horas
    search_query = request.GET.get('search', '')
    alertas = [
        {'maquina': maquina, 'horas_restantes': maquina.horas_restantes_mantenimiento()}
        for maquina in Maquina.objects.filter(nombre__icontains=search_query)
        if maquina.horas_restantes_mantenimiento() <= horas_alerta
    ]
    context = {
        'alertas': sorted(alertas, key=lambda x: x['horas_restantes']),
        'total_alertas': len(alertas),
        'alert_form': alert_form,
    }
    return render(request, 'mUP_maquina/alertas.html', context)

@login_required
def tabla_mantenimientos(request):
    maquinas = Maquina.objects.all()
    tipos_mantenimiento = TipoMantenimientoMaquina.objects.all()
    for maquina in maquinas:
        maquina.mantenimientos = maquina.mantenimientomaquina_set.all().order_by('-fecha_fin', '-hora_fin')
    context = {
        'maquinas': maquinas,
        'tipos_mantenimiento': tipos_mantenimiento,
    }
    return render(request, 'mUP_maquina/tablas.html', context)

@login_required
def crear_maquina(request):
    form = MaquinaForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        intervalo_mantenimiento = form.cleaned_data['intervalo_mantenimiento']
        if intervalo_mantenimiento < 0:
            form.add_error('intervalo_mantenimiento', 'El intervalo de mantenimiento no puede ser un número negativo')
        else:
            if 'imagen' in request.FILES:
                form.instance.imagen = request.FILES['imagen']
            form.save()
            return redirect('maquina')
        messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
    context = {'form': form}
    return render(request, 'mUP_maquina/nueva.html', context)

@login_required    
def detalles(request, id):
    maquina = get_object_or_404(Maquina, id=id)
    mantenimientos = maquina.mantenimientomaquina_set.all().order_by('-fecha_fin', '-hora_fin')
    
    if request.method == 'POST':
        form = MaquinaForm(request.POST, request.FILES, instance=maquina)
        if form.is_valid():
            intervalo_mantenimiento = form.cleaned_data.get('intervalo_mantenimiento')
            if intervalo_mantenimiento < 0:
                form.add_error('intervalo_mantenimiento', 'El intervalo de mantenimiento no puede ser un número negativo')
            else:
                form.save()
                return render(request, 'mUP_maquina/detalles.html', {
                    'maquina': maquina,
                    'form': form,
                    'id': id,
                    'mantenimientos': mantenimientos
                })
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

    form = MaquinaForm(instance=maquina)
    return render(request, 'mUP_maquina/detalles.html', {
        'maquina': maquina,
        'form': form,
        'id': id,
        'mantenimientos': mantenimientos
    })

@login_required
def eliminar(request, id):
    maquina = get_object_or_404(Maquina, id=id)
    maquina.delete()
    return redirect ('maquina') 

# fin de vistas generales----------------------------------------------------------------------------------


@login_required
def eliminar_mantenimiento(request, id):
    mantenimiento = get_object_or_404(MantenimientoMaquina, id=id)
    mantenimiento.delete()
    previous_url = request.META.get('HTTP_REFERER')
    return HttpResponseRedirect(previous_url)

@login_required
def mantenimientos_maquina_preventivo(request, id):
    maquina = get_object_or_404(Maquina, id=id)
    tipo_mantenimiento = get_object_or_404(TipoMantenimientoMaquina, id=2) 
    mantenimientos = maquina.mantenimientomaquina_set.filter(tipo=tipo_mantenimiento).order_by('-fecha_fin', '-hora_fin')
    context = {
        'maquina': maquina,
        'tipo_mantenimiento': tipo_mantenimiento,
        'mantenimientos': mantenimientos,
    }
    return render(request, 'mUP_maquina/manteniminetos_preventivo.html', context)   

@login_required
def mod_mantenimiento_maquina_preventivo(request, id):
    mantenimiento = get_object_or_404(MantenimientoMaquina, id=id)
    maquina = mantenimiento.maquina
    tipo_mantenimiento = get_object_or_404(TipoMantenimientoMaquina, id=2)
    
    if request.method == 'GET':
        form_mant = MantenimientoMaquinaPreventivoForm(instance=mantenimiento)
        return render(request, 'mUP_maquina/mod_mantenimineto_preventivo.html', 
                     {'form_mant': form_mant, 'maquina': maquina})
    
    if request.method == 'POST':
        form_mant = MantenimientoMaquinaPreventivoForm(
            request.POST, 
            request.FILES, 
            instance=mantenimiento
        )
        if not form_mant.is_valid():
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
            return render(request, 'mUP_maquina/mod_mantenimineto_preventivo.html',
                         {'form_mant': form_mant, 'maquina': maquina})

        if form_mant.cleaned_data['hr_maquina'] > maquina.horas_máquina_trabajada:
            form_mant.add_error('hr_maquina', 
                'Las horas de trabajo del mantenimiento no pueden ser mayores que las horas de trabajo de la máquina.')
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
            return render(request, 'mUP_maquina/nuevo_mantenimineto_preventivo.html',
                         {'form_mant': form_mant, 'maquina': maquina, 'tipo_mantenimiento': tipo_mantenimiento})
        mantenimiento = form_mant.save(commit=False)
        mantenimiento.maquina = maquina
        mantenimiento.tipo = tipo_mantenimiento
        mantenimiento.partes_y_piezas = ""
        if 'imagen' in request.FILES:
            mantenimiento.imagen = request.FILES['imagen']
        mantenimiento.save()
        return redirect('mantenimientos_maquina_preventivo', id=maquina.id)
    return HttpResponse("Method Not Allowed", status=405)

@login_required
def nuevo_mantenimiento_maquina_preventivo(request, id):
    maquina = get_object_or_404(Maquina, id=id)
    tipo_mantenimiento = get_object_or_404(TipoMantenimientoMaquina, id=2)
    
    if request.method == 'GET':
        form_mant = MantenimientoMaquinaPreventivoForm()
        return render(request, 'mUP_maquina/nuevo_mantenimineto_preventivo.html',
                     {'form_mant': form_mant, 'maquina': maquina, 'tipo_mantenimiento': tipo_mantenimiento})
    
    if request.method == 'POST':
        form_mant = MantenimientoMaquinaPreventivoForm(request.POST, request.FILES)
        if not form_mant.is_valid():
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
            return render(request, 'mUP_maquina/nuevo_mantenimineto_preventivo.html',
                         {'form_mant': form_mant, 'maquina': maquina, 'tipo_mantenimiento': tipo_mantenimiento})
        if form_mant.cleaned_data['hr_maquina'] > maquina.horas_máquina_trabajada:
            form_mant.add_error('hr_maquina', 
                'Las horas de trabajo del mantenimiento no pueden ser mayores que las horas de trabajo de la máquina.')
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
            return render(request, 'mUP_maquina/nuevo_mantenimineto_preventivo.html',
                         {'form_mant': form_mant, 'maquina': maquina, 'tipo_mantenimiento': tipo_mantenimiento})
        mantenimiento = form_mant.save(commit=False)
        mantenimiento.maquina = maquina
        mantenimiento.tipo = tipo_mantenimiento
        mantenimiento.partes_y_piezas = ""
        if 'imagen' in request.FILES:
            mantenimiento.imagen = request.FILES['imagen']
        mantenimiento.save()
        return redirect('mantenimientos_maquina_preventivo', id=maquina.id)
    return HttpResponse("Method Not Allowed", status=405)

@login_required
def mantenimientos_maquina_correctivo(request, id):
    maquina = get_object_or_404(Maquina, id=id)
    tipo_mantenimiento = get_object_or_404(TipoMantenimientoMaquina, id=1) 
    mantenimientos = maquina.mantenimientomaquina_set.filter(tipo=tipo_mantenimiento).order_by('-fecha_fin', '-hora_fin')
    context = {
        'maquina': maquina,
        'tipo_mantenimiento': tipo_mantenimiento,
        'mantenimientos': mantenimientos,
    }
    return render(request, 'mUP_maquina/manteniminetos_correctivo.html', context)   

@login_required
def mod_mantenimiento_maquina_correctivo(request, id):
    mantenimiento = get_object_or_404(MantenimientoMaquina, id=id)
    maquina = mantenimiento.maquina
    tipo_mantenimiento = get_object_or_404(TipoMantenimientoMaquina, id=1)
    
    if request.method == 'GET':
        form_mant = MantenimientoMaquinaCorrectivoForm(instance=mantenimiento)
        return render(request, 'mUP_maquina/mod_mantenimineto_correctivo.html',
                     {'form_mant': form_mant, 'maquina': maquina})
    
    if request.method == 'POST':
        form_mant = MantenimientoMaquinaCorrectivoForm(
            request.POST, 
            request.FILES, 
            instance=mantenimiento
        )
        if not form_mant.is_valid():
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
            return render(request, 'mUP_maquina/mod_mantenimineto_correctivo.html',
                         {'form_mant': form_mant, 'maquina': maquina})
        if form_mant.cleaned_data['hr_maquina'] > maquina.horas_máquina_trabajada:
            form_mant.add_error('hr_maquina', 
                'Las horas de trabajo del mantenimiento no pueden ser mayores que las horas de trabajo de la máquina.')
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
            return render(request, 'mUP_maquina/nuevo_mantenimineto_correctivo.html',
                         {'form_mant': form_mant, 'maquina': maquina, 'tipo_mantenimiento': tipo_mantenimiento})
        mantenimiento = form_mant.save(commit=False)
        mantenimiento.maquina = maquina
        mantenimiento.tipo = tipo_mantenimiento
        if 'imagen' in request.FILES:
            mantenimiento.imagen = request.FILES['imagen']
        mantenimiento.save()
        return redirect('mantenimientos_maquina_correctivo', id=maquina.id)
    return HttpResponse("Method Not Allowed", status=405)

@login_required
def nuevo_mantenimineto_maquina_correctivo(request, id):
    maquina = get_object_or_404(Maquina, id=id)
    tipo_mantenimiento = get_object_or_404(TipoMantenimientoMaquina, id=1)
    
    if request.method == 'GET':
        form_mant = MantenimientoMaquinaCorrectivoForm()
        return render(request, 'mUP_maquina/nuevo_mantenimineto_correctivo.html',
                     {'form_mant': form_mant, 'maquina': maquina, 'tipo_mantenimiento': tipo_mantenimiento})
    
    if request.method == 'POST':
        form_mant = MantenimientoMaquinaCorrectivoForm(request.POST, request.FILES)

        if not form_mant.is_valid():
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
            return render(request, 'mUP_maquina/nuevo_mantenimineto_correctivo.html',
                         {'form_mant': form_mant, 'maquina': maquina, 'tipo_mantenimiento': tipo_mantenimiento})
        if form_mant.cleaned_data['hr_maquina'] > maquina.horas_máquina_trabajada:
            form_mant.add_error('hr_maquina', 
                'Las horas de trabajo del mantenimiento no pueden ser mayores que las horas de trabajo de la máquina.')
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
            return render(request, 'mUP_maquina/nuevo_mantenimineto_correctivo.html',
                         {'form_mant': form_mant, 'maquina': maquina, 'tipo_mantenimiento': tipo_mantenimiento})
        mantenimiento = form_mant.save(commit=False)
        mantenimiento.maquina = maquina
        mantenimiento.tipo = tipo_mantenimiento
        if 'imagen' in request.FILES:
            mantenimiento.imagen = request.FILES['imagen']    
        mantenimiento.save()
        return redirect('mantenimientos_maquina_correctivo', id=maquina.id)
    return HttpResponse("Method Not Allowed", status=405)


# descargas -----------------------------------------------------------------------------------

@login_required
def documento_general_mantenimientos_maquina(request):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    tipo_mantenimiento_id = request.GET.get('tipo_mantenimiento')

    mantenimientos = MantenimientoMaquina.objects.filter(fecha_fin__year=anio)

    if mes:
        mantenimientos = mantenimientos.filter(fecha_fin__month=mes)

    if tipo_mantenimiento_id:  # Si se seleccionó un tipo de mantenimiento
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoMaquina, pk=tipo_mantenimiento_id)
        mantenimientos = mantenimientos.filter(tipo=tipo_mantenimiento)

    mantenimientos = mantenimientos.order_by('-fecha_fin', '-hora_fin')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if mes:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_maquinas_{}_{}.xlsx"'.format(mes, anio)
    else:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_maquinas_{}.xlsx"'.format(anio)

    wb = openpyxl.Workbook()
    ws = wb.active

    headers = ['Maquina', 'Tipo', 'Operador', 'Fecha I', 'Hora I', 'Fecha F', 'Hora F', 'Hr Máquina', 'Partes y Piezas', 'Descripción']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    row = 2
    for mantenimiento in mantenimientos:
        ws.append([
            mantenimiento.maquina.nombre,
            mantenimiento.tipo.tipo,
            mantenimiento.operador,
            mantenimiento.fecha_inicio,
            mantenimiento.hora_inicio,
            mantenimiento.fecha_fin,
            mantenimiento.hora_fin,
            mantenimiento.hr_maquina,
            mantenimiento.partes_y_piezas,
            mantenimiento.descripción
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response




@login_required
def documento_mantenimientos_preventivos_maquina(request, id):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    tipo_mantenimiento_id = 2

    maquina = get_object_or_404(Maquina, pk=id)
    mantenimientos = MantenimientoMaquina.objects.filter(maquina=maquina).order_by('-fecha_fin', '-hora_fin')

    if mes:
        mantenimientos = mantenimientos.filter(fecha_fin__month=mes)
    if anio:
        mantenimientos = mantenimientos.filter(fecha_fin__year=anio)
    if tipo_mantenimiento_id: # Si se seleccionó un tipo de mantenimiento
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoMaquina, pk=tipo_mantenimiento_id)
        mantenimientos = mantenimientos.filter(tipo=tipo_mantenimiento)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if mes:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_preventivos_de_{}_{}_{}.xlsx"'.format(maquina.nombre, mes, anio)
    else:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_preventivos_de_{}_{}.xlsx"'.format(maquina.nombre, anio)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Define los encabezados de la tabla
    headers = ['Operador', 'Fecha I', 'Hora I', 'Fecha F', 'Hora F', 'Hr Máquina', 'Descripción']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    # Agrega los datos de los mantenimientos
    row = 2
    for mantenimiento in mantenimientos:
        ws.append([
            mantenimiento.operador,
            mantenimiento.fecha_inicio,
            mantenimiento.hora_inicio,
            mantenimiento.fecha_fin,
            mantenimiento.hora_fin,
            mantenimiento.hr_maquina,
            mantenimiento.descripción
        ])

    # Ajusta el ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response




@login_required
def documento_mantenimientos_correctivos_maquina(request, id):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    tipo_mantenimiento_id = 1

    maquina = get_object_or_404(Maquina, pk=id)
    mantenimientos = MantenimientoMaquina.objects.filter(maquina=maquina).order_by('-fecha_fin', '-hora_fin')

    if mes:
        mantenimientos = mantenimientos.filter(fecha_fin__month=mes)
    if anio:
        mantenimientos = mantenimientos.filter(fecha_fin__year=anio)
    if tipo_mantenimiento_id: # Si se seleccionó un tipo de mantenimiento
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoMaquina, pk=tipo_mantenimiento_id)
        mantenimientos = mantenimientos.filter(tipo=tipo_mantenimiento)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if mes:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_correctivos_de_{}_{}_{}.xlsx"'.format(maquina.nombre, mes, anio)
    else:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_correctivos_de_{}_{}.xlsx"'.format(maquina.nombre, anio)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Define los encabezados de la tabla
    headers = ['Operador', 'Fecha I', 'Hora I', 'Fecha F', 'Hora F', 'Hr Máquina', 'Partes y Piezas', 'Descripción']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    # Agrega los datos de los mantenimientos
    row = 2
    for mantenimiento in mantenimientos:
        ws.append([
            mantenimiento.operador,
            mantenimiento.fecha_inicio,
            mantenimiento.hora_inicio,
            mantenimiento.fecha_fin,
            mantenimiento.hora_fin,
            mantenimiento.hr_maquina,
            mantenimiento.partes_y_piezas,
            mantenimiento.descripción
        ])

    # Ajusta el ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response