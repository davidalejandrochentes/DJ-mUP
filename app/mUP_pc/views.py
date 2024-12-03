from django.shortcuts import render, redirect, get_object_or_404
from .models import PC, MantenimientoPC, TipoMantenimientoPC, DiasParaAlerta
from .forms import PCForm, MantenimientoPCForm, DiasParaAlertaForm
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.http import HttpResponse, HttpResponseRedirect
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage

import openpyxl
from openpyxl.styles import Font, PatternFill



# vistas generales ----------------------------------------------------------

@login_required
def pc(request):
    search_query = request.GET.get('search', '')
    dias_alerta = DiasParaAlerta.objects.first().días
    pcs_list = PC.objects.filter(nombre__icontains=search_query)
    
    alertas = []
    for pc in PC.objects.all():
        dias_restantes = pc.dias_restantes_mantenimiento()
        if dias_restantes <= dias_alerta:
            alertas.append({
                'pc': pc,
                'dias_restantes': dias_restantes
            })
    alertas_ordenadas = sorted(alertas, key=lambda x: x['dias_restantes'])
    
    # Implementar paginación
    page = request.GET.get('page', 1)
    paginator = Paginator(pcs_list, 10)  # 10 PCs por página
    try:
        pcs = paginator.page(page)
    except PageNotAnInteger:
        pcs = paginator.page(1)
    except EmptyPage:
        pcs = paginator.page(paginator.num_pages)
    
    return render(request, 'mUP_pc/pc.html', {
        'pcs': pcs,
        'total_pcs': pcs_list.count(),
        'alertas': alertas_ordenadas,
        'total_alertas': len(alertas_ordenadas),
    })

@login_required
def alertas(request):
    dias_alert = get_object_or_404(DiasParaAlerta, id=1)
    search_query = request.GET.get('search', '')
    if request.method == 'POST':
        alert_form = DiasParaAlertaForm(request.POST, instance=dias_alert)
        if alert_form.is_valid():
            días = alert_form.cleaned_data.get('días')
            if días >= 1:
                alert_form.save()
            return redirect('pc_alertas')
    else:
        alert_form = DiasParaAlertaForm(instance=dias_alert)
    
    alertas = []
    for pc in PC.objects.filter(nombre__icontains=search_query):
        dias_restantes = pc.dias_restantes_mantenimiento()
        if dias_restantes <= dias_alert.días:
            alertas.append({
                'pc': pc,
                'dias_restantes': dias_restantes
            })
    alertas_ordenadas = sorted(alertas, key=lambda x: x['dias_restantes'])
    
    # Implementar paginación
    page = request.GET.get('page', 1)
    paginator = Paginator(alertas_ordenadas, 10)  # 10 alertas por página
    try:
        alertas_paginadas = paginator.page(page)
    except PageNotAnInteger:
        alertas_paginadas = paginator.page(1)
    except EmptyPage:
        alertas_paginadas = paginator.page(paginator.num_pages)
    
    return render(request, 'mUP_pc/alertas.html', {
        'alertas': alertas_paginadas,
        'total_alertas': len(alertas_ordenadas),
        'alert_form': alert_form,
    })

@login_required
def tabla_mantenimientos(request):
    pcs_list = PC.objects.prefetch_related('mantenimientopc_set').all()
    for pc in pcs_list:
        pc.mantenimientos = pc.mantenimientopc_set.order_by('-fecha_fin', '-hora_fin')
    
    # Implementar paginación
    page = request.GET.get('page', 1)
    paginator = Paginator(pcs_list, 10)  # 10 PCs por página
    try:
        pcs = paginator.page(page)
    except PageNotAnInteger:
        pcs = paginator.page(1)
    except EmptyPage:
        pcs = paginator.page(paginator.num_pages)
    
    context = {
        'pcs': pcs,
        'total_pcs': pcs_list.count(),
        'tipos_mantenimiento': TipoMantenimientoPC.objects.all(),
    }
    return render(request, 'mUP_pc/tablas.html', context)

# vistas de creación --------------------------------------------------------

@login_required
def crear_pc(request):
    if request.method == 'GET':
        return render(request, 'mUP_pc/nueva.html', {'form': PCForm()})
    
    if request.method == 'POST':
        form = PCForm(request.POST, request.FILES)
        if not form.is_valid():
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
            return render(request, 'mUP_pc/nueva.html', {'form': form})   
        intervalo_mantenimiento = form.cleaned_data.get('intervalo_mantenimiento')
        if intervalo_mantenimiento < 0:
            form.add_error('intervalo_mantenimiento', 'El intervalo de mantenimiento no puede ser un número negativo')
            return render(request, 'mUP_pc/nueva.html', {'form': form})
        if 'imagen' in request.FILES:
            form.instance.imagen = request.FILES['imagen']
        form.save()
        return redirect('pc')

# vistas de detalles --------------------------------------------------------

@login_required    
def detalles(request, id):
    pc = get_object_or_404(PC, id=id)
    mantenimientos = pc.mantenimientopc_set.all().order_by('-fecha_fin', '-hora_fin')

    if request.method == 'GET':
        return render(request, 'mUP_pc/detalles.html', {
            'pc': pc,
            'form': PCForm(instance=pc),
            'id': id,
            'mantenimientos': mantenimientos,
        })
    
    if request.method == 'POST':
        form = PCForm(request.POST, request.FILES, instance=pc)
        if not form.is_valid():
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))   
        intervalo_mantenimiento = form.cleaned_data.get('intervalo_mantenimiento')
        if intervalo_mantenimiento < 0:
            form.add_error('intervalo_mantenimiento', 'El intervalo de mantenimiento no puede ser un número negativo')
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))   
        form.save()
        return render(request, 'mUP_pc/detalles.html', {
            'pc': pc,
            'form': form,
            'id': id,
            'mantenimientos': mantenimientos,
        })

# vistas de eliminación -----------------------------------------------------

@login_required
def eliminar(request, id):
    pc = get_object_or_404(PC, id=id)
    pc.delete()
    return redirect('pc')

# fin de vistas generales---------------------------------------------    




# vistas de mantenimientos --------------------------------------------------

@login_required
def eliminar_mantenimiento(request, id):
    mantenimiento = get_object_or_404(MantenimientoPC, id=id)
    mantenimiento.delete()
    previous_url = request.META.get('HTTP_REFERER')
    return HttpResponseRedirect(previous_url)

@login_required
def mantenimientos_pc(request, id, mant):
    pc = get_object_or_404(PC, id=id)
    tipo_mantenimiento = get_object_or_404(TipoMantenimientoPC, id=mant) 
    mantenimientos_list = pc.mantenimientopc_set.filter(tipo=tipo_mantenimiento).order_by('-fecha_fin', '-hora_fin')
    
    # Implementar paginación
    page = request.GET.get('page', 1)
    paginator = Paginator(mantenimientos_list, 10)  # 10 mantenimientos por página
    try:
        mantenimientos = paginator.page(page)
    except PageNotAnInteger:
        mantenimientos = paginator.page(1)
    except EmptyPage:
        mantenimientos = paginator.page(paginator.num_pages)
    
    context = {
        'pc': pc,
        'tipo_mantenimiento': tipo_mantenimiento,
        'mantenimientos': mantenimientos,
        'total_mantenimientos': mantenimientos_list.count(),
    }
    return render(request, 'mUP_pc/mantenimientos_pc.html', context)

@login_required
def mod_mantenimiento_pc(request, id, mant):
    mantenimiento = get_object_or_404(MantenimientoPC, id=id)
    tipo_mantenimiento = get_object_or_404(TipoMantenimientoPC, id=mant)
    pc = mantenimiento.pc

    if request.method == 'GET':
        form_mant = MantenimientoPCForm(instance=mantenimiento)
    elif request.method == 'POST':
        form_mant = MantenimientoPCForm(request.POST, request.FILES, instance=mantenimiento)
        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.pc = pc
            mantenimiento.tipo = tipo_mantenimiento
            mantenimiento.save()
            return redirect('mantenimientos_pc', id=pc.id, mant=mant)
        messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")

    context = {
        'form_mant': form_mant,
        'pc': pc,
        'tipo_mantenimiento': tipo_mantenimiento
    }
    return render(request, 'mUP_pc/mod_mantenimiento.html', context) if request.method in ['GET', 'POST'] else HttpResponse("Method Not Allowed", status=405)

@login_required
def nuevo_mantenimiento_pc(request, id, mant):
    pc = get_object_or_404(PC, id=id)
    tipo_mantenimiento = get_object_or_404(TipoMantenimientoPC, id=mant)

    if request.method == 'GET':
        form_mant = MantenimientoPCForm()
    elif request.method == 'POST':
        form_mant = MantenimientoPCForm(request.POST, request.FILES)
        if form_mant.is_valid():
            mantenimiento = form_mant.save(commit=False)
            mantenimiento.pc = pc
            mantenimiento.tipo = tipo_mantenimiento
            mantenimiento.save()
            return redirect('mantenimientos_pc', id=pc.id, mant=mant)
        messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")

    context = {
        'form_mant': form_mant,
        'pc': pc,
        'tipo_mantenimiento': tipo_mantenimiento,
    }
    return render(request, 'mUP_pc/nuevo_mantenimiento.html', context) if request.method in ['GET', 'POST'] else HttpResponse("Method Not Allowed", status=405)







@login_required
def documento_general_mantenimientos_pc(request):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    tipo_mantenimiento_id = request.GET.get('tipo_mantenimiento')

    mantenimientos = MantenimientoPC.objects.filter(fecha_fin__year=anio)

    if mes:
        mantenimientos = mantenimientos.filter(fecha_fin__month=mes)
    if tipo_mantenimiento_id:  # Si se seleccionó un tipo de mantenimiento
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoPC, pk=tipo_mantenimiento_id)
        mantenimientos = mantenimientos.filter(tipo=tipo_mantenimiento)

    mantenimientos = mantenimientos.order_by('-fecha_fin', '-hora_fin')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if mes:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_equipos_de_cómputo_{}_{}.xlsx"'.format(mes, anio)
    else:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_equipos_de_cómputo_{}.xlsx"'.format(anio)

    wb = openpyxl.Workbook()
    ws = wb.active

    headers = ['E Cómputo', 'Tipo', 'Fecha I', 'Hora I', 'Fecha F', 'Hora F', 'Operador', 'Partes y Piezas', 'Descripción']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    row = 2
    for mantenimiento in mantenimientos:
        ws.append([
            mantenimiento.pc.nombre,
            mantenimiento.tipo.tipo,
            mantenimiento.fecha_inicio,
            mantenimiento.hora_inicio,
            mantenimiento.fecha_fin,
            mantenimiento.hora_fin,
            mantenimiento.operador,
            mantenimiento.partes_y_piezas,
            mantenimiento.descripción
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response




@login_required
def documento_mantenimientos_pc(request, id, mant):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    tipo_mantenimiento_id = mant
    tipo_mantenimiento = get_object_or_404(TipoMantenimientoPC, id=mant) 

    pc = get_object_or_404(PC, pk=id)
    mantenimientos = MantenimientoPC.objects.filter(pc=pc).order_by('-fecha_fin', '-hora_fin')

    if mes:
        mantenimientos = mantenimientos.filter(fecha_fin__month=mes)
    if anio:
        mantenimientos = mantenimientos.filter(fecha_fin__year=anio)
    if tipo_mantenimiento_id: # Si se seleccionó un tipo de mantenimiento
        tipo_mantenimiento = get_object_or_404(TipoMantenimientoPC, pk=tipo_mantenimiento_id)
        mantenimientos = mantenimientos.filter(tipo=tipo_mantenimiento)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if mes:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_{}_de_{}_{}_{}.xlsx"'.format(tipo_mantenimiento.tipo, pc.nombre, mes, anio)
    else:
        response['Content-Disposition'] = 'attachment; filename="mantenimientos_{}_de_{}_{}.xlsx"'.format(tipo_mantenimiento.tipo, pc.nombre, anio)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Define los encabezados de la tabla
    headers = ['Fecha I', 'Hora I', 'Fecha F', 'Hora F', 'Operador', 'Partes y Piezas', 'Descripción']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    # Agrega los datos de los mantenimientos
    row = 2
    for mantenimiento in mantenimientos:
        ws.append([
            mantenimiento.fecha_inicio,
            mantenimiento.hora_inicio,
            mantenimiento.fecha_fin,
            mantenimiento.hora_fin,
            mantenimiento.operador,
            mantenimiento.partes_y_piezas,
            mantenimiento.descripción
        ])

    # Ajusta el ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response